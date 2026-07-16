import sqlite3
import json
from openpyxl import Workbook

db = "../23240175-seeding.db"
output_file = "23240175-sq26-results.xlsx"
keywords_json = "isic_keywords_map.json"


def load_tax():
    with open(keywords_json, encoding="utf-8") as f:
        return json.load(f)


def rank_classes(text, tax):
    if not text:
        return []
    t = text.lower()
    hits = {}
    for code, info in tax.items():
        s = sum(t.count(kw) for kw in info["keywords"] if kw)
        if s > 0:
            hits[code] = s
    return sorted(hits.items(), key=lambda x: x[1], reverse=True)


def clean_code(raw):
    return raw.replace("-", "") if raw else ""


def main():
    tax = load_tax()
    conn = sqlite3.connect(db)
    conn.row_factory = sqlite3.Row

    projects = conn.execute("""
        SELECT id, repository_id, type, title, class, description
        FROM projects
        ORDER BY repository_id, id
    """).fetchall()

    print(f"Number of projects: {len(projects)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "repository_id",
        "project_type",
        "project_title",
        "primary_class",
        "secondary_class",
        "no_project_files",
    ]
    # Write header row
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    row = 2
    for proj in projects:
        pid = proj["id"]

        n = conn.execute(
            "SELECT COUNT(1) FROM files WHERE project_id=?", (pid,)
        ).fetchone()[0]

        primary = clean_code(proj["class"]) if proj["class"] else ""

        secondary = ""
        if proj["class"]:
            kw_rows = conn.execute(
                "SELECT keyword FROM keywords WHERE project_id=?", (pid,)
            ).fetchall()
            kw_text = " ".join(k[0] for k in kw_rows if k[0])
            full_text = " ".join(filter(None, [proj["title"], proj["description"], kw_text]))
            ranked = rank_classes(full_text, tax)
            if len(ranked) > 1:
                secondary = clean_code(ranked[1][0])

        ws.cell(row=row, column=1, value=proj["repository_id"])
        ws.cell(row=row, column=2, value=proj["type"])
        ws.cell(row=row, column=3, value=proj["title"])
        ws.cell(row=row, column=4, value=primary)
        ws.cell(row=row, column=5, value=secondary)
        ws.cell(row=row, column=6, value=n)
        row += 1

    ws.freeze_panes = "A2"
    wb.save(output_file)
    conn.close()

    print(f"\nSaved: {output_file}")
    print(f"Rows: {len(projects)}")


if __name__ == "__main__":
    main()